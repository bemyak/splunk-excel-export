# -*- coding: utf-8 -*-

from collections import deque
import json
import logging
import openpyxl as openpyxl
import datetime
from openpyxl.styles import Font, PatternFill, Color, Border, Side

logger = logging.getLogger('splunk')

FLUSH_COUNT = 1000
MAX_ROWS = 65000

thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

header_size = 14
iso_fmt = "%Y-%m-%dT%H:%M:%S.%f"
fmt = "%Y-%m-%d %H:%M:%S"


class WBClient(object):
    """ consumer class for asycore client callback """

    def __init__(self, template=None):
        self.decoder = json.JSONDecoder()
        self.input_buffer = deque()
        self.output_buffer = []
        self.result_count = 0
        self.touched = False

        if not template:
            self.wb = openpyxl.Workbook()
        else:
            self.wb = openpyxl.load_workbook(filename=template)
        self.wb.active = 0

    def add_data(self, json_objects, field_names, date_from=None, date_to=None):
        if not self.touched:
            self.touched = True
        else:
            self.next_sheet()

        self.write_header(date_from, date_to, 1)

        current_row = 3
        """
        adds data to the current wb sheet
        """
        ws = self.wb.active
        for i, field_name in enumerate(field_names):
            cell = ws.cell(row=current_row, column=i+1)
            cell.value = field_name
            cell.fill = PatternFill(fgColor=Color(
                'E7E6E6'), patternType='solid')
            cell.font = Font(bold=True, name='Arial')
            cell.border = thin_border
        current_row += 1

        for row, (json_object, _) in enumerate(json_objects):
            for col, field in enumerate(field_names):
                tmp = json_object['result'].get(field)
                if tmp:
                    if isinstance(tmp, list):
                        tmp = unicode(','.join(tmp))
                    else:
                        tmp = unicode(tmp)

                cell = ws.cell(row=current_row+row, column=col+1)
                cell.value = self.cast_value(tmp)
                cell.border = thin_border
            if current_row+row >= MAX_ROWS:
                break

        self.set_filters()
        self.set_width()

    def next_sheet(self):
        """
        increment the sheet index
        set result_sheet and write header
        """
        prev_index = self.wb._active_sheet_index
        next_index = prev_index + 1
        self.wb.active = next_index
        if self.wb.active is None:
            self.wb.create_sheet("Sheet" + str(next_index))

    def save(self, filename):
        self.wb.active = 0
        self.wb.save(filename)

    def cast_value(self, value):
        if value is None:
            return None
        try:
            return int(value)
        except ValueError:
            pass
        try:
            return float(value)
        except ValueError:
            pass
        return self.max_len(value)

    def set_filters(self):
        ws = self.wb.active
        isFirst = True
        first = ''
        last = ''
        for cells in ws.iter_cols(min_row=3):
            for cell in cells:
                if isFirst:
                    first = cell.coordinate
                    isFirst = False
                last = cell.coordinate
        cell_range = ':'.join([first, last])
        ws.auto_filter.ref = cell_range

    def set_width(self):
        ws = self.wb.active
        for col in ws.columns:
            max_length = 0
            column = col[0].column
            for cell in col:
                if cell.coordinate in ws.merged_cells:
                    continue
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width

    def write_header(self, date_from=None, date_to=None, row=1):
        """
        writes the excel header
        """
        date_from = self.format_date(date_from)
        if date_from is None:
            date_from = 'from the beginning'
        date_to = self.format_date(date_to)
        if date_to is None:
            date_to = 'till now'
        ws = self.wb.active
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
        a1 = ws.cell(column=1, row=row)
        a1.font = Font(size=header_size, bold=True, name='Arial')
        a1.value = u'Report start date: ' + date_from + u'; end date: ' + date_to
        rd = ws.row_dimensions[row]
        rd.height = header_size + 2

    def format_date(self, date):
        if date is None:
            return None
        try:
            date = datetime.datetime.strptime(date[:-6], iso_fmt)
        except Exception, ex:
            logger.exception(ex)
            return None
        if (date.year == 1970):
            return None
        return date.strftime(fmt)

    def max_len(self, data):
        """
        excel does now allow any cell to be > 32767 characters
        """
        if len(data) > 32767:
            logger.warn('reducing to 32767 characters')
            data = data[:32766]
        return data
